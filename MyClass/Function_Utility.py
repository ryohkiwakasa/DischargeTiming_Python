import numpy as np
import copy
import pyvista as pv
import openpyxl as px


def read_csv_outputdata(excel_filepath, sheet_name, nozzle_num):
    print("hello")


def output_env_param_excel(array, filename, ws_title="test"):
    book = px.Workbook()
    ws = book.worksheets[0]
    ws.title = ws_title
    ws.append(["gap distance", "incidence angle", "head angle"])
    for i in range(array.shape[0]):
        ws.append(array[i, :].tolist())
    book.save(filename)
    print("save complete!")

def eliminate_error_value(array, n_mov=10):
    output = copy.copy(array)
    std3 = np.std(array)
    mean_val = np.mean(array)
    indexes = np.where(abs(array-mean_val) > std3)[0]
    for idx in indexes:
        t_idx = np.where(abs(array[idx-n_mov:idx+n_mov]) < std3)[0]
        print("idx: ", idx)
        print("t_idx: ", t_idx)
        output[idx] = np.mean(array[t_idx])
    return output


def output_timing_data(TimingArray, filepath, big_or_small=True):
    print("--- output_timing_data--- @Function Utility")
    genre = 0 if big_or_small==True else 0

    book = px.Workbook()
    book.create_sheet()
    ws_on_off = book.worksheets[0]
    ws_delay = book.worksheets[1]
    ws_on_off.title = "onoff"
    ws_delay.title = "delay"

    ws_on_off.append([genre])
    ws_on_off.append([int(np.max(TimingArray[0].sub_label))])
    ws_delay.append([genre])
    ws_delay.append([int(np.max(TimingArray[0].sub_label))])

    for label in range(int(np.max(TimingArray[0].sub_label))):
        i = 1
        for dpiTiming in TimingArray:
            index = np.where(dpiTiming.sub_label == label + 1)[0]
            target_on_off = dpiTiming.on_off[index]
            target_delay = dpiTiming.delay_time[index]
            print("label: ", label + 1, "nozzle", i, "data num: ", target_on_off.shape[0])
            # print("target_on_off: ", target_on_off[100:150], "\n")
            if i == 1:
                ws_on_off.append([target_on_off.shape[0]])
                ws_delay.append([target_on_off.shape[0]])
            ws_on_off.append(target_on_off.tolist())
            ws_delay.append(target_delay.tolist())
            i += 1
    book.save(filepath)
    print("save complete!\n")

def calculation_opentime(input_param, param_table, param1, param2, param3):
    if np.any(param1 == input_param[0]):
        index_1 = np.where(param1 == input_param[0])[0][0]
    else:
        index_1 = np.searchsorted(param1, input_param[0], side='left') - 1

    if np.any(param2 == input_param[1]):
        index_2 = np.where(param2 == input_param[1])[0][0]
    else:
        index_2 = np.searchsorted(param2, input_param[1], side='left') - 1

    if np.any(param3 == input_param[2]):
        index_3 = np.where(param3 == input_param[2])[0][0]
    else:
        index_3 = np.searchsorted(param3, input_param[2], side='left') - 1

    rate_1 = (param1[index_1+1] - input_param[0]) / (param1[index_1+1] - param1[index_1])
    rate_2 = (param2[index_2+1] - input_param[1]) / (param2[index_2+1] - param2[index_2])
    rate_3 = (param3[index_3+1] - input_param[2]) / (param3[index_3+1] - param3[index_3])

    tmp1 = param_table[index_1, :, :] * rate_1 + param_table[index_1+1, :, :] * (1 - rate_1)
    tmp2 = tmp1[index_2, :] * rate_2 + tmp1[index_2+1, :] * (1 - rate_2)
    answer = tmp2[index_3] * rate_3 + tmp2[index_3+1] * (1 - rate_3)

    return answer


